import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook

# Configuracion para Spark en Windows
if os.name == "nt":
    hadoop_home = os.environ.get("HADOOP_HOME", r"C:\hadoop")
    if os.path.isdir(hadoop_home):
        os.environ["HADOOP_HOME"] = hadoop_home
        hadoop_bin = os.path.join(hadoop_home, "bin")
        if os.path.isdir(hadoop_bin):
            os.environ["PATH"] = hadoop_bin + os.pathsep + os.environ.get("PATH", "")

from pyspark.sql import SparkSession
from pyspark.sql import functions as F

BASE_DIR = Path(os.environ.get("PROYECTO_BASE_DIR", Path(__file__).resolve().parent)).resolve()
PARQUET_PATH = str(BASE_DIR / "consolidado_produccion_investigadores.parquet")
OUTPUT_XLSX = str(BASE_DIR / "resumen_validacion_union.xlsx")

# Columnas solicitadas por el usuario (sin duplicar NME_NIV_FORM_PR)
requested_cols = [
    "ID_PERSONA_PD",
    "NME_GENERO_PR",
    "NME_PAIS_NAC_PR",
    "NME_REGION_NAC_PR",
    "NME_DEPARTAMENTO_NAC_PR",
    "NME_MUNICIPIO_NAC_PR",
    "NME_NIV_FORM_PR",
    "NME_CLASIFICACION_PR",
    "EDAD_ANOS_PR",
    "NME_DEPARTAMENTO_RES_PR",
    "NME_REGION_RES_PR",
    "NME_PAIS_RES_PR",
    "INST_FILIA",
    "ID_VICTIMA_CONFLICTO",
    "TXT_GRUPO_ETNICO",
    "NME_GRAN_AREA_PR",
    "TXT_POBLACION_DISCA",
    "NME_TIPO_MEDICION_PD",
    "NME_TIPOLOGIA_PD",
    "ID_TIPO_PD_MED",
    "NME_CATEGORIA_PD",
    "ID_CONVOCATORIA",
    "NME_CONVOCATORIA",
    "ANO_CONVO",
    "ID_PERSONA_PR",
    "COD_GRUPO_GR",
    # — Grupo de investigación —
    "INST_AVAL",
    "NME_CLASIFICACION_GR",
    "NME_PAIS_GR",
    "NME_REGION_GR",
    "NME_DEPARTAMENTO_GR",
    "NME_MUNICIPIO_GR",
]

spark = (
    SparkSession.builder
    .appName("Resumen_Validacion_Union")
    .config("spark.driver.memory", "4g")
    .config("spark.executor.memory", "4g")
    .config("spark.sql.shuffle.partitions", "50")
    .getOrCreate()
)
spark.sparkContext.setLogLevel("WARN")

df = spark.read.parquet(PARQUET_PATH)
total_rows = df.count()
all_cols = set(df.columns)

# Hoja 1: verificación de columnas solicitadas
presence_rows = []
for c in requested_cols:
    presence_rows.append({
        "columna": c,
        "presente_en_union": "SI" if c in all_cols else "NO"
    })
df_presence = pd.DataFrame(presence_rows)

# Trabajar solo con columnas existentes
existing_cols = [c for c in requested_cols if c in all_cols]
df_sel = df.select(*existing_cols)

# Hoja 2: resumen estadístico por columna
agg_exprs = []
for c in existing_cols:
    agg_exprs.append(F.sum(F.when(F.col(c).isNull() | (F.trim(F.col(c)) == ""), 1).otherwise(0)).alias(f"nulls__{c}"))
    agg_exprs.append(F.approx_count_distinct(F.col(c)).alias(f"distinct__{c}"))

agg_row = df_sel.agg(*agg_exprs).collect()[0].asDict()

summary_rows = []
for c in existing_cols:
    nulls = int(agg_row.get(f"nulls__{c}", 0) or 0)
    distincts = int(agg_row.get(f"distinct__{c}", 0) or 0)
    non_nulls = int(total_rows - nulls)
    null_pct = round((nulls / total_rows) * 100, 4) if total_rows else 0.0

    summary_rows.append({
        "columna": c,
        "filas_totales": int(total_rows),
        "no_nulos": non_nulls,
        "nulos_o_vacios": nulls,
        "porcentaje_nulos": null_pct,
        "valores_distintos_aprox": distincts,
    })

df_summary = pd.DataFrame(summary_rows)

total_cols_union = len(df.columns)

# Hoja adicional: resumen general de filas/columnas de la union
df_overview = pd.DataFrame([
    {"metrica": "filas_totales_union", "valor": int(total_rows)},
    {"metrica": "columnas_totales_union", "valor": int(total_cols_union)},
    {"metrica": "columnas_solicitadas", "valor": int(len(requested_cols))},
    {"metrica": "columnas_solicitadas_presentes", "valor": int(len(existing_cols))},
])

# Hoja adicional: listado de todas las columnas del archivo unido
df_columns_union = pd.DataFrame({"columna_union": df.columns})

# Hoja adicional: muestra amplia para inspeccion visual de la union
sample_rows = 160000
df_sample = df_sel.limit(sample_rows).toPandas()

# Crear workbook en modo streaming para alto volumen de filas
wb = Workbook(write_only=True)

# Hoja verificacion_columnas
ws_presence = wb.create_sheet(title="verificacion_columnas")
ws_presence.append(list(df_presence.columns))
for row_values in df_presence.itertuples(index=False, name=None):
    ws_presence.append(list(row_values))

# Hoja resumen_columnas
ws_summary = wb.create_sheet(title="resumen_columnas")
ws_summary.append(list(df_summary.columns))
for row_values in df_summary.itertuples(index=False, name=None):
    ws_summary.append(list(row_values))

# Hoja filas y columnas de la union
ws_overview = wb.create_sheet(title="filas_columnas_union")
ws_overview.append(list(df_overview.columns))
for row_values in df_overview.itertuples(index=False, name=None):
    ws_overview.append(list(row_values))

# Hoja listado de columnas existentes en la union
ws_columns_union = wb.create_sheet(title="columnas_union")
ws_columns_union.append(list(df_columns_union.columns))
for row_values in df_columns_union.itertuples(index=False, name=None):
    ws_columns_union.append(list(row_values))

# Hoja de muestra de datos
ws_sample = wb.create_sheet(title="muestra_160000")
ws_sample.append(list(df_sample.columns))
for row_values in df_sample.itertuples(index=False, name=None):
    ws_sample.append(list(row_values))

wb.save(OUTPUT_XLSX)

spark.stop()

print(f"Excel generado: {OUTPUT_XLSX}")
print(f"Filas en union: {total_rows}")
print(f"Columnas en union: {total_cols_union}")
print(f"Filas exportadas (muestra): {sample_rows}")
print(f"Columnas solicitadas: {len(requested_cols)}")
print(f"Columnas presentes: {len(existing_cols)}")
