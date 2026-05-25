import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook

# Configuracion para Spark en Windows
if os.name == "nt":
    hadoop_home = os.environ.get("HADOOP_HOME", r"C:\hadoop")
    if os.path.isdir(hadoop_home):
        os.environ["HADOOP_HOME"] = hadoop_home
        hadoop_bin = os.path.join(hadoop_home, "bin")
        if os.path.isdir(hadoop_bin):
            os.environ["PATH"] = hadoop_bin + os.pathsep + os.environ.get("PATH", "")

from pyspark.sql import SparkSession
from pyspark.sql import functions as F


BASE_DIR = Path(os.environ.get("PROYECTO_BASE_DIR", Path(__file__).resolve().parent)).resolve()
PARQUET_PATH = str(BASE_DIR / "consolidado_produccion_investigadores_match.parquet")
OUTPUT_XLSX = str(BASE_DIR / "revision_visualizacion_match.xlsx")

MATCH_COLUMNS = [
    "INST_FILIA",
    "NME_GENERO_PR",
    "NME_PAIS_NAC_PR",
    "NME_REGION_NAC_PR",
    "NME_DEPARTAMENTO_NAC_PR",
    "NME_MUNICIPIO_NAC_PR",
    "NME_NIV_FORM_PR",
    "NME_CLASIFICACION_PR",
    "EDAD_ANOS_PR",
    "NME_DEPARTAMENTO_RES_PR",
    "NME_REGION_RES_PR",
    "NME_PAIS_RES_PR",
    "ID_VICTIMA_CONFLICTO",
    "TXT_GRUPO_ETNICO",
    "NME_GRAN_AREA_PR",
    "TXT_POBLACION_DISCA",
]


def invalid_condition(column_name):
    value = F.trim(F.coalesce(F.col(column_name).cast("string"), F.lit("")))
    return (value == "") | (value == "0")


spark = (
    SparkSession.builder
    .appName("Revision_Visualizacion_Match")
    .config("spark.driver.memory", "4g")
    .config("spark.executor.memory", "4g")
    .config("spark.sql.shuffle.partitions", "50")
    .getOrCreate()
)
spark.sparkContext.setLogLevel("WARN")

df = spark.read.parquet(PARQUET_PATH)
total_rows = df.count()
total_cols = len(df.columns)

presence_rows = []
for column in MATCH_COLUMNS:
    presence_rows.append({
        "columna": column,
        "presente_en_parquet": "SI" if column in df.columns else "NO",
    })
df_presence = pd.DataFrame(presence_rows)

agg_exprs = []
for column in MATCH_COLUMNS:
    agg_exprs.append(F.sum(F.when(F.col(column).isNull(), 1).otherwise(0)).alias(f"null__{column}"))
    agg_exprs.append(F.sum(F.when(invalid_condition(column), 1).otherwise(0)).alias(f"empty_or_zero__{column}"))

agg_row = df.agg(*agg_exprs).collect()[0].asDict()

validation_rows = []
for column in MATCH_COLUMNS:
    nulls = int(agg_row.get(f"null__{column}", 0) or 0)
    empty_or_zero = int(agg_row.get(f"empty_or_zero__{column}", 0) or 0)
    invalids = nulls + empty_or_zero
    validation_rows.append({
        "columna": column,
        "nulos": nulls,
        "vacios_o_cero": empty_or_zero,
        "invalidos_totales": invalids,
        "estado": "PASS" if invalids == 0 else "FAIL",
    })
df_validation = pd.DataFrame(validation_rows)

summary_rows = [
    {"metrica": "filas_totales_match", "valor": int(total_rows)},
    {"metrica": "columnas_totales_match", "valor": int(total_cols)},
    {"metrica": "columnas_criterio_match", "valor": int(len(MATCH_COLUMNS))},
    {"metrica": "columnas_validadas_ok", "valor": int((df_validation["estado"] == "PASS").sum())},
]
df_summary = pd.DataFrame(summary_rows)

df_columns = pd.DataFrame({"columna_parquet": df.columns})

sample_rows = min(5000, total_rows)
df_sample = df.limit(sample_rows).toPandas()

wb = Workbook(write_only=True)

ws_summary = wb.create_sheet(title="resumen_general")
ws_summary.append(list(df_summary.columns))
for row_values in df_summary.itertuples(index=False, name=None):
    ws_summary.append(list(row_values))

ws_presence = wb.create_sheet(title="columnas_match")
ws_presence.append(list(df_presence.columns))
for row_values in df_presence.itertuples(index=False, name=None):
    ws_presence.append(list(row_values))

ws_validation = wb.create_sheet(title="validacion_match")
ws_validation.append(list(df_validation.columns))
for row_values in df_validation.itertuples(index=False, name=None):
    ws_validation.append(list(row_values))

ws_columns = wb.create_sheet(title="columnas_parquet")
ws_columns.append(list(df_columns.columns))
for row_values in df_columns.itertuples(index=False, name=None):
    ws_columns.append(list(row_values))

ws_sample = wb.create_sheet(title="muestra_real")
ws_sample.append(list(df_sample.columns))
for row_values in df_sample.itertuples(index=False, name=None):
    ws_sample.append(list(row_values))

wb.save(OUTPUT_XLSX)

spark.stop()

print(f"Excel generado: {OUTPUT_XLSX}")
print(f"Filas en parquet match: {total_rows}")
print(f"Columnas en parquet match: {total_cols}")
print(f"Filas exportadas en muestra: {sample_rows}")
print(f"Validacion PASS: {int((df_validation['estado'] == 'PASS').sum())}/{len(MATCH_COLUMNS)}")